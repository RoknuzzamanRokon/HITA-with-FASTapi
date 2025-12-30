"""
Export Format Handler Service

Handles transformation of data into various export formats:
- CSV: Flattened structure with UTF-8 BOM encoding
- JSON: Nested structure with metadata
- Excel: Multi-sheet workbook with formatting
"""

import csv
import json
import io
import logging
from typing import Generator, List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path

from models import Hotel, Location, Contact, ProviderMapping, SupplierSummary
from export_schemas import ExportMetadata

# Configure logging
logger = logging.getLogger(__name__)


class ExportFormatHandler:
    """
    Service for transforming data into export file formats.
    
    Supports:
    - CSV with UTF-8 BOM encoding and flattened structure
    - JSON with nested structure and metadata
    - Excel with multiple sheets and formatting
    """
    
    def __init__(self):
        """Initialize ExportFormatHandler"""
        logger.info("ExportFormatHandler initialized")

    def to_csv(
        self,
        data: Generator[List[Any], None, None],
        output_path: str,
        headers: List[str],
        flatten_func: callable = None
    ) -> str:
        """
        Generate CSV file from data stream with UTF-8 BOM encoding.
        
        Features:
        - UTF-8 encoding with BOM for Excel compatibility
        - Streaming generation for memory efficiency
        - Batch processing support
        - Proper escaping of special characters
        
        Args:
            data: Generator yielding batches of records
            output_path: Path where CSV file should be written
            headers: List of column headers
            flatten_func: Optional function to flatten complex objects
            
        Returns:
            Path to the generated CSV file
        """
        logger.info(f"Generating CSV export to {output_path}")
        logger.debug(f"CSV headers: {headers}")
        
        try:
            # Ensure output directory exists
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Open file with UTF-8 BOM encoding
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.DictWriter(
                    csvfile,
                    fieldnames=headers,
                    quoting=csv.QUOTE_MINIMAL,
                    escapechar='\\'
                )
                
                # Write header row
                writer.writeheader()
                logger.debug("CSV header row written")
                
                # Process data in batches
                total_rows = 0
                for batch in data:
                    for record in batch:
                        # Flatten record if function provided
                        if flatten_func:
                            flattened = flatten_func(record)
                        elif isinstance(record, dict):
                            flattened = record
                        else:
                            # Assume it's a model object, convert to dict
                            flattened = self._model_to_dict(record)
                        
                        # Ensure all headers are present
                        row = {header: flattened.get(header, '') for header in headers}
                        
                        # Write row
                        writer.writerow(row)
                        total_rows += 1
                    
                    if total_rows % 1000 == 0:
                        logger.debug(f"Written {total_rows} rows to CSV")
            
            logger.info(f"CSV export completed: {total_rows} rows written to {output_path}")
            return output_path
            
        except IOError as e:
            logger.error(f"File I/O error generating CSV export: {str(e)}")
            raise IOError(f"Failed to write CSV file: {str(e)}")
        except Exception as e:
            logger.error(f"Error generating CSV export: {str(e)}")
            raise Exception(f"CSV generation failed: {str(e)}")

    def flatten_hotel_data(self, hotel: Hotel) -> Dict[str, Any]:
        """
        Flatten nested hotel structure for CSV export.
        
        Combines data from:
        - Hotel base fields
        - First location (primary location)
        - Contacts (concatenated)
        - Provider mappings (concatenated)
        
        Args:
            hotel: Hotel model object
            
        Returns:
            Flattened dictionary with all data in single level
        """
        logger.debug(f"Flattening hotel data for ITTID: {hotel.ittid}")
        
        flattened = {
            'ittid': hotel.ittid or '',
            'name': hotel.name or '',
            'latitude': hotel.latitude or '',
            'longitude': hotel.longitude or '',
            'rating': hotel.rating or '',
            'address_line1': hotel.address_line1 or '',
            'address_line2': hotel.address_line2 or '',
            'postal_code': hotel.postal_code or '',
            'property_type': hotel.property_type or '',
            'primary_photo': hotel.primary_photo or '',
            'created_at': hotel.created_at.isoformat() if hotel.created_at else '',
            'updated_at': hotel.updated_at.isoformat() if hotel.updated_at else '',
        }
        
        # Add location data (use first location if multiple)
        if hotel.locations and len(hotel.locations) > 0:
            location = hotel.locations[0]
            flattened.update({
                'city_name': location.city_name or '',
                'state_name': location.state_name or '',
                'state_code': location.state_code or '',
                'country_name': location.country_name or '',
                'country_code': location.country_code or '',
                'master_city_name': location.master_city_name or '',
                'city_code': location.city_code or '',
                'city_location_id': location.city_location_id or '',
            })
        else:
            flattened.update({
                'city_name': '',
                'state_name': '',
                'state_code': '',
                'country_name': '',
                'country_code': '',
                'master_city_name': '',
                'city_code': '',
                'city_location_id': '',
            })
        
        # Add contacts (concatenate multiple contacts)
        if hotel.contacts:
            phones = [c.value for c in hotel.contacts if c.contact_type == 'phone']
            emails = [c.value for c in hotel.contacts if c.contact_type == 'email']
            websites = [c.value for c in hotel.contacts if c.contact_type == 'website']
            faxes = [c.value for c in hotel.contacts if c.contact_type == 'fax']
            
            flattened.update({
                'phone': '; '.join(phones) if phones else '',
                'email': '; '.join(emails) if emails else '',
                'website': '; '.join(websites) if websites else '',
                'fax': '; '.join(faxes) if faxes else '',
            })
        else:
            flattened.update({
                'phone': '',
                'email': '',
                'website': '',
                'fax': '',
            })
        
        # Add provider mappings (concatenate)
        if hotel.provider_mappings:
            provider_names = [pm.provider_name for pm in hotel.provider_mappings]
            provider_ids = [pm.provider_id for pm in hotel.provider_mappings]
            giata_codes = [pm.giata_code for pm in hotel.provider_mappings if pm.giata_code]
            vervotech_ids = [pm.vervotech_id for pm in hotel.provider_mappings if pm.vervotech_id]
            
            flattened.update({
                'provider_names': '; '.join(provider_names) if provider_names else '',
                'provider_ids': '; '.join(provider_ids) if provider_ids else '',
                'giata_codes': '; '.join(giata_codes) if giata_codes else '',
                'vervotech_ids': '; '.join(vervotech_ids) if vervotech_ids else '',
            })
        else:
            flattened.update({
                'provider_names': '',
                'provider_ids': '',
                'giata_codes': '',
                'vervotech_ids': '',
            })
        
        return flattened

    def _model_to_dict(self, model: Any) -> Dict[str, Any]:
        """
        Convert SQLAlchemy model to dictionary.
        
        Args:
            model: SQLAlchemy model object
            
        Returns:
            Dictionary representation of model
        """
        if hasattr(model, '__dict__'):
            result = {}
            for key, value in model.__dict__.items():
                if not key.startswith('_'):
                    if isinstance(value, datetime):
                        result[key] = value.isoformat()
                    else:
                        result[key] = value
            return result
        return {}

    def get_csv_headers_hotel(self) -> List[str]:
        """
        Get standard CSV headers for hotel export.
        
        Returns:
            List of column headers
        """
        return [
            'ittid', 'name', 'latitude', 'longitude', 'rating',
            'address_line1', 'address_line2', 'postal_code', 'property_type',
            'city_name', 'state_name', 'state_code', 'country_name', 'country_code',
            'master_city_name', 'city_code', 'city_location_id',
            'phone', 'email', 'website', 'fax',
            'provider_names', 'provider_ids', 'giata_codes', 'vervotech_ids',
            'primary_photo', 'created_at', 'updated_at'
        ]

    def get_content_type(self, format: str) -> str:
        """
        Get HTTP Content-Type header for export format.
        
        Args:
            format: Export format (csv, json, excel)
            
        Returns:
            Content-Type header value
        """
        content_types = {
            'csv': 'text/csv; charset=utf-8',
            'json': 'application/json; charset=utf-8',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        return content_types.get(format.lower(), 'application/octet-stream')

    def get_filename(self, export_type: str, format: str, timestamp: datetime = None) -> str:
        """
        Generate filename for export file.
        
        Args:
            export_type: Type of export (hotels, mappings, supplier_summary)
            format: Export format (csv, json, excel)
            timestamp: Optional timestamp for filename
            
        Returns:
            Filename with extension
        """
        if timestamp is None:
            timestamp = datetime.utcnow()
        
        timestamp_str = timestamp.strftime('%Y%m%d_%H%M%S')
        extensions = {
            'csv': 'csv',
            'json': 'json',
            'excel': 'xlsx'
        }
        ext = extensions.get(format.lower(), 'dat')
        
        return f"{export_type}_{timestamp_str}.{ext}"

    def to_json(
        self,
        data: Generator[List[Any], None, None],
        output_path: str,
        metadata: Optional[ExportMetadata] = None,
        preserve_structure: bool = True
    ) -> str:
        """
        Generate JSON file from data stream with metadata.
        
        Features:
        - Nested structure preservation for relationships
        - Export metadata header
        - ISO 8601 datetime formatting
        - Pretty-printing with 2-space indentation
        - Streaming support for large datasets
        
        Args:
            data: Generator yielding batches of records
            output_path: Path where JSON file should be written
            metadata: Optional export metadata to include in file
            preserve_structure: Whether to preserve nested relationships
            
        Returns:
            Path to the generated JSON file
        """
        logger.info(f"Generating JSON export to {output_path}")
        
        try:
            # Ensure output directory exists
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Collect all records (JSON needs full structure)
            all_records = []
            total_records = 0
            
            for batch in data:
                for record in batch:
                    if preserve_structure:
                        json_record = self._model_to_json_dict(record)
                    elif isinstance(record, dict):
                        json_record = record
                    else:
                        json_record = self._model_to_dict(record)
                    
                    all_records.append(json_record)
                    total_records += 1
                
                if total_records % 1000 == 0:
                    logger.debug(f"Processed {total_records} records for JSON export")
            
            # Build final JSON structure
            json_output = {}
            
            # Add metadata if provided
            if metadata:
                json_output['metadata'] = {
                    'export_id': metadata.export_id,
                    'generated_at': metadata.generated_at.isoformat(),
                    'generated_by': metadata.generated_by,
                    'user_id': metadata.user_id,
                    'filters_applied': metadata.filters_applied,
                    'total_records': len(all_records),
                    'format': metadata.format,
                    'version': metadata.version
                }
            
            # Add data
            json_output['data'] = all_records
            json_output['record_count'] = len(all_records)
            
            # Write to file with pretty-printing
            with open(output_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(json_output, jsonfile, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"JSON export completed: {total_records} records written to {output_path}")
            return output_path
            
        except IOError as e:
            logger.error(f"File I/O error generating JSON export: {str(e)}")
            raise IOError(f"Failed to write JSON file: {str(e)}")
        except Exception as e:
            logger.error(f"Error generating JSON export: {str(e)}")
            raise Exception(f"JSON generation failed: {str(e)}")

    def _model_to_json_dict(self, model: Any) -> Dict[str, Any]:
        """
        Convert SQLAlchemy model to JSON-friendly dictionary with nested relationships.
        
        Preserves relationships like locations, contacts, and provider_mappings.
        
        Args:
            model: SQLAlchemy model object
            
        Returns:
            Dictionary with nested structure
        """
        if isinstance(model, Hotel):
            return self._hotel_to_json_dict(model)
        elif isinstance(model, ProviderMapping):
            return self._mapping_to_json_dict(model)
        elif isinstance(model, SupplierSummary):
            return self._supplier_summary_to_json_dict(model)
        else:
            # Generic conversion
            return self._model_to_dict(model)

    def _hotel_to_json_dict(self, hotel: Hotel) -> Dict[str, Any]:
        """
        Convert Hotel model to JSON dictionary with nested relationships.
        
        Args:
            hotel: Hotel model object
            
        Returns:
            Dictionary with nested locations, contacts, and mappings
        """
        hotel_dict = {
            'ittid': hotel.ittid,
            'name': hotel.name,
            'latitude': hotel.latitude,
            'longitude': hotel.longitude,
            'rating': hotel.rating,
            'address_line1': hotel.address_line1,
            'address_line2': hotel.address_line2,
            'postal_code': hotel.postal_code,
            'property_type': hotel.property_type,
            'primary_photo': hotel.primary_photo,
            'created_at': hotel.created_at.isoformat() if hotel.created_at else None,
            'updated_at': hotel.updated_at.isoformat() if hotel.updated_at else None,
        }
        
        # Add locations
        if hotel.locations:
            hotel_dict['locations'] = [
                {
                    'city_name': loc.city_name,
                    'state_name': loc.state_name,
                    'state_code': loc.state_code,
                    'country_name': loc.country_name,
                    'country_code': loc.country_code,
                    'master_city_name': loc.master_city_name,
                    'city_code': loc.city_code,
                    'city_location_id': loc.city_location_id,
                }
                for loc in hotel.locations
            ]
        
        # Add contacts
        if hotel.contacts:
            hotel_dict['contacts'] = [
                {
                    'contact_type': contact.contact_type,
                    'value': contact.value,
                }
                for contact in hotel.contacts
            ]
        
        # Add provider mappings
        if hotel.provider_mappings:
            hotel_dict['provider_mappings'] = [
                {
                    'provider_name': pm.provider_name,
                    'provider_id': pm.provider_id,
                    'system_type': pm.system_type,
                    'vervotech_id': pm.vervotech_id,
                    'giata_code': pm.giata_code,
                    'created_at': pm.created_at.isoformat() if pm.created_at else None,
                    'updated_at': pm.updated_at.isoformat() if pm.updated_at else None,
                }
                for pm in hotel.provider_mappings
            ]
        
        return hotel_dict

    def _mapping_to_json_dict(self, mapping: ProviderMapping) -> Dict[str, Any]:
        """
        Convert ProviderMapping model to JSON dictionary.
        
        Args:
            mapping: ProviderMapping model object
            
        Returns:
            Dictionary with mapping data
        """
        mapping_dict = {
            'ittid': mapping.ittid,
            'provider_name': mapping.provider_name,
            'provider_id': mapping.provider_id,
            'created_at': mapping.created_at.isoformat() if mapping.created_at else None,
            'updated_at': mapping.updated_at.isoformat() if mapping.updated_at else None,
        }
        
        # Add hotel info if available
        if mapping.hotel:
            mapping_dict['hotel'] = {
                'name': mapping.hotel.name,
                'latitude': mapping.hotel.latitude,
                'longitude': mapping.hotel.longitude,
                'rating': mapping.hotel.rating,
                'country_code': mapping.hotel.locations[0].country_code if mapping.hotel.locations else None,
            }
        
        return mapping_dict

    def _supplier_summary_to_json_dict(self, summary: SupplierSummary) -> Dict[str, Any]:
        """
        Convert SupplierSummary model to JSON dictionary.
        
        Args:
            summary: SupplierSummary model object
            
        Returns:
            Dictionary with supplier summary data
        """
        return {
            'provider_name': summary.provider_name,
            'total_hotels': summary.total_hotels,
            'total_mappings': summary.total_mappings,
            'last_updated': summary.last_updated.isoformat() if summary.last_updated else None,
            'summary_generated_at': summary.summary_generated_at.isoformat() if summary.summary_generated_at else None,
        }

    def to_excel(
        self,
        data: Dict[str, Generator[List[Any], None, None]],
        output_path: str,
        metadata: Optional[ExportMetadata] = None
    ) -> str:
        """
        Generate Excel file with multiple sheets and formatting.
        
        Features:
        - Multi-sheet structure (Hotels, Locations, Contacts, Mappings, Summary)
        - Formatted headers with bold text and colored background
        - Auto-column sizing
        - Freeze panes on header rows
        - Summary dashboard sheet with key metrics
        
        Args:
            data: Dictionary mapping sheet names to data generators
            output_path: Path where Excel file should be written
            metadata: Optional export metadata to include
            
        Returns:
            Path to the generated Excel file
        """
        logger.info(f"Generating Excel export to {output_path}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Ensure output directory exists
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Process each sheet
            for sheet_name, data_generator in data.items():
                logger.debug(f"Creating sheet: {sheet_name}")
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Collect data for this sheet
                sheet_data = []
                for batch in data_generator:
                    sheet_data.extend(batch)
                
                if not sheet_data:
                    logger.warning(f"No data for sheet {sheet_name}")
                    continue
                
                # Determine headers based on first record
                if isinstance(sheet_data[0], dict):
                    headers = list(sheet_data[0].keys())
                else:
                    # Convert model to dict to get headers
                    first_dict = self._model_to_dict(sheet_data[0])
                    headers = list(first_dict.keys())
                
                # Write headers with formatting
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data rows
                for row_idx, record in enumerate(sheet_data, start=2):
                    if isinstance(record, dict):
                        row_data = record
                    else:
                        row_data = self._model_to_dict(record)
                    
                    for col_idx, header in enumerate(headers, start=1):
                        value = row_data.get(header, '')
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-size columns
                for col_idx, header in enumerate(headers, start=1):
                    column_letter = get_column_letter(col_idx)
                    max_length = len(str(header))
                    
                    # Check first 100 rows for max length
                    for row_idx in range(2, min(102, len(sheet_data) + 2)):
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    
                    # Set column width (max 50 characters)
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
                
                # Freeze header row
                ws.freeze_panes = ws['A2']
                
                logger.debug(f"Sheet {sheet_name} completed with {len(sheet_data)} rows")
            
            # Add summary sheet if metadata provided
            if metadata:
                self._add_summary_sheet(wb, metadata)
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Excel export completed: {output_path}")
            return output_path
            
        except ImportError:
            logger.error("openpyxl library not installed. Cannot generate Excel export.")
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        except IOError as e:
            logger.error(f"File I/O error generating Excel export: {str(e)}")
            raise IOError(f"Failed to write Excel file: {str(e)}")
        except Exception as e:
            logger.error(f"Error generating Excel export: {str(e)}")
            raise Exception(f"Excel generation failed: {str(e)}")

    def _add_summary_sheet(self, workbook: Any, metadata: ExportMetadata) -> None:
        """
        Add summary dashboard sheet to Excel workbook.
        
        Args:
            workbook: openpyxl Workbook object
            metadata: Export metadata
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            logger.debug("Adding summary dashboard sheet")
            
            # Create summary sheet at the beginning
            ws = workbook.create_sheet(title="Summary", index=0)
            
            # Title
            ws['A1'] = "Export Summary"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:B1')
            
            # Metadata
            row = 3
            summary_data = [
                ("Export ID", metadata.export_id),
                ("Generated At", metadata.generated_at.strftime('%Y-%m-%d %H:%M:%S')),
                ("Generated By", metadata.generated_by),
                ("User ID", metadata.user_id),
                ("Total Records", metadata.total_records),
                ("Format", metadata.format),
                ("Version", metadata.version),
            ]
            
            for label, value in summary_data:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            # Filters applied
            row += 1
            ws.cell(row=row, column=1, value="Filters Applied").font = Font(bold=True, size=12)
            row += 1
            
            for filter_key, filter_value in metadata.filters_applied.items():
                if filter_value:
                    ws.cell(row=row, column=1, value=filter_key)
                    ws.cell(row=row, column=2, value=str(filter_value))
                    row += 1
            
            # Auto-size columns
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 50
            
            logger.debug("Summary sheet added successfully")
            
        except Exception as e:
            logger.error(f"Error adding summary sheet: {str(e)}")
            # Don't fail the entire export if summary fails
            pass

    def to_excel_hotels(
        self,
        hotels: Generator[List[Hotel], None, None],
        output_path: str,
        metadata: Optional[ExportMetadata] = None
    ) -> str:
        """
        Generate Excel file for hotel export with multiple sheets.
        
        Creates sheets:
        - Hotels: Main hotel data
        - Locations: Location details
        - Contacts: Contact information
        - Mappings: Provider mappings
        - Summary: Export metadata and statistics
        
        Args:
            hotels: Generator yielding batches of Hotel objects
            output_path: Path where Excel file should be written
            metadata: Optional export metadata
            
        Returns:
            Path to the generated Excel file
        """
        logger.info(f"Generating multi-sheet Excel export for hotels to {output_path}")
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Ensure output directory exists
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            
            # Collect all hotel data
            all_hotels = []
            for batch in hotels:
                all_hotels.extend(batch)
            
            logger.debug(f"Collected {len(all_hotels)} hotels for Excel export")
            
            # Create workbook
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Sheet 1: Hotels
            self._create_hotels_sheet(wb, all_hotels)
            
            # Sheet 2: Locations
            self._create_locations_sheet(wb, all_hotels)
            
            # Sheet 3: Contacts
            self._create_contacts_sheet(wb, all_hotels)
            
            # Sheet 4: Mappings
            self._create_mappings_sheet(wb, all_hotels)
            
            # Sheet 5: Summary
            if metadata:
                self._add_summary_sheet(wb, metadata)
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Multi-sheet Excel export completed: {output_path}")
            return output_path
            
        except ImportError:
            logger.error("openpyxl library not installed. Cannot generate Excel export.")
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        except Exception as e:
            logger.error(f"Error generating Excel export: {str(e)}")
            raise

    def _create_hotels_sheet(self, workbook: Any, hotels: List[Hotel]) -> None:
        """Create Hotels sheet in Excel workbook"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title="Hotels")
        
        headers = ['ITTID', 'Name', 'Latitude', 'Longitude', 'Rating', 'Address Line 1', 
                   'Address Line 2', 'Postal Code', 'Property Type', 'Primary Photo', 
                   'Created At', 'Updated At']
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, hotel in enumerate(hotels, start=2):
            ws.cell(row=row_idx, column=1, value=hotel.ittid)
            ws.cell(row=row_idx, column=2, value=hotel.name)
            ws.cell(row=row_idx, column=3, value=hotel.latitude)
            ws.cell(row=row_idx, column=4, value=hotel.longitude)
            ws.cell(row=row_idx, column=5, value=hotel.rating)
            ws.cell(row=row_idx, column=6, value=hotel.address_line1)
            ws.cell(row=row_idx, column=7, value=hotel.address_line2)
            ws.cell(row=row_idx, column=8, value=hotel.postal_code)
            ws.cell(row=row_idx, column=9, value=hotel.property_type)
            ws.cell(row=row_idx, column=10, value=hotel.primary_photo)
            ws.cell(row=row_idx, column=11, value=hotel.created_at.strftime('%Y-%m-%d %H:%M:%S') if hotel.created_at else '')
            ws.cell(row=row_idx, column=12, value=hotel.updated_at.strftime('%Y-%m-%d %H:%M:%S') if hotel.updated_at else '')
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        ws.freeze_panes = ws['A2']

    def _create_locations_sheet(self, workbook: Any, hotels: List[Hotel]) -> None:
        """Create Locations sheet in Excel workbook"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title="Locations")
        
        headers = ['ITTID', 'City Name', 'State Name', 'State Code', 'Country Name', 
                   'Country Code', 'Master City Name', 'City Code', 'City Location ID']
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        row_idx = 2
        for hotel in hotels:
            if hotel.locations:
                for location in hotel.locations:
                    ws.cell(row=row_idx, column=1, value=hotel.ittid)
                    ws.cell(row=row_idx, column=2, value=location.city_name)
                    ws.cell(row=row_idx, column=3, value=location.state_name)
                    ws.cell(row=row_idx, column=4, value=location.state_code)
                    ws.cell(row=row_idx, column=5, value=location.country_name)
                    ws.cell(row=row_idx, column=6, value=location.country_code)
                    ws.cell(row=row_idx, column=7, value=location.master_city_name)
                    ws.cell(row=row_idx, column=8, value=location.city_code)
                    ws.cell(row=row_idx, column=9, value=location.city_location_id)
                    row_idx += 1
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        ws.freeze_panes = ws['A2']

    def _create_contacts_sheet(self, workbook: Any, hotels: List[Hotel]) -> None:
        """Create Contacts sheet in Excel workbook"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title="Contacts")
        
        headers = ['ITTID', 'Hotel Name', 'Contact Type', 'Value']
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        row_idx = 2
        for hotel in hotels:
            if hotel.contacts:
                for contact in hotel.contacts:
                    ws.cell(row=row_idx, column=1, value=hotel.ittid)
                    ws.cell(row=row_idx, column=2, value=hotel.name)
                    ws.cell(row=row_idx, column=3, value=contact.contact_type)
                    ws.cell(row=row_idx, column=4, value=contact.value)
                    row_idx += 1
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        ws.freeze_panes = ws['A2']

    def _create_mappings_sheet(self, workbook: Any, hotels: List[Hotel]) -> None:
        """Create Mappings sheet in Excel workbook"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        ws = workbook.create_sheet(title="Mappings")
        
        headers = ['ITTID', 'Hotel Name', 'Provider Name', 'Provider ID', 'System Type', 
                   'Vervotech ID', 'Giata Code', 'Created At', 'Updated At']
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        row_idx = 2
        for hotel in hotels:
            if hotel.provider_mappings:
                for mapping in hotel.provider_mappings:
                    ws.cell(row=row_idx, column=1, value=hotel.ittid)
                    ws.cell(row=row_idx, column=2, value=hotel.name)
                    ws.cell(row=row_idx, column=3, value=mapping.provider_name)
                    ws.cell(row=row_idx, column=4, value=mapping.provider_id)
                    ws.cell(row=row_idx, column=5, value=mapping.system_type)
                    ws.cell(row=row_idx, column=6, value=mapping.vervotech_id)
                    ws.cell(row=row_idx, column=7, value=mapping.giata_code)
                    ws.cell(row=row_idx, column=8, value=mapping.created_at.strftime('%Y-%m-%d %H:%M:%S') if mapping.created_at else '')
                    ws.cell(row=row_idx, column=9, value=mapping.updated_at.strftime('%Y-%m-%d %H:%M:%S') if mapping.updated_at else '')
                    row_idx += 1
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        ws.freeze_panes = ws['A2']
